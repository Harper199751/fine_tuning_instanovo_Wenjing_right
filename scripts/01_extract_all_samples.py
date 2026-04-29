import os  # handle file paths and folders
import re #extract scan numbers using regular expressions
from collections import defaultdict #set a space for each raw files
from openpyxl import load_workbook #read .xlsx

#set pathway
excel_file = r"d:\01 PHD\03 Courses\CIE500 WENJING\CIE500\0.project\scan_list.xlsx"
mgf_folder = r"d:\01 PHD\03 Courses\CIE500 WENJING\CIE500\0.project"
output_folder = r"d:\01 PHD\03 Courses\CIE500 WENJING\CIE500\0.project\filtered_output"

#merge .mgf
merge_all = True # merge all mgfs
merged_output_file = os.path.join(output_folder, "all_filtered_spectra.mgf") # merged all mgfs in a subfolder
os.makedirs(output_folder, exist_ok=True) #run many times without debug

# read .xlsx
wb = load_workbook(excel_file, read_only=True, data_only=True) #set the input format of my .xlsx 
ws = wb.active # use the first worksheet

rows = list(ws.iter_rows(values_only=True)) #read and save row value
headers = list(rows[0])  #read and save row headers
print("Excel headers:", headers)

raw_file_col = headers.index("Raw file") #find the column of "Raw file"
scan_number_col = headers.index("Scan number") #find the column of "scan number"

# build a dictionary to save each scan number according to their raw file, otherwise is the blank 
scan_dict = defaultdict(set)

#save the set of scans for each sample
for row in rows[1:]:
    raw_file = row[raw_file_col]
    scan_number = row[scan_number_col]

    if raw_file is not None and scan_number is not None:
        scan_dict[str(raw_file)].add(int(scan_number))

print(f"Found{len(scan_dict)} sample") 

# clean old file if run this py again
if merge_all:
    with open(merged_output_file, "w", encoding="utf-8") as f:
        pass

# take out the scan according to raw sample
for sample_name, target_scans in scan_dict.items():
    mgf_file = os.path.join(mgf_folder, sample_name + ".mgf") #find mathced .mgf
    output_file = os.path.join(output_folder, sample_name + "_filtered.mgf") #output name
#check the files are exist
#skip the unfound .mgf
    if not os.path.exists(mgf_file):
       print(f"[Skipped] File not found: {mgf_file}")
       continue

    print(f"\nProcessing: {sample_name}")
    print(f"Number of target scans: {len(target_scans)}")

    kept_count = 0 #save the number of scans in one raw file
    current_block = [] # save current scan information
    inside_block = False 

    # read .mgf
    with open(mgf_file, "r", encoding="utf-8", errors="ignore") as fin, \
         open(output_file, "w", encoding="utf-8") as fout:

        # read by lines from BEGIN IONS to END IONS
        for line in fin:
            line_strip = line.strip()

            if line_strip == "BEGIN IONS":
                inside_block = True
                current_block = [line]

            elif line_strip == "END IONS":
                current_block.append(line)
                inside_block = False

                scan_number = None
                for block_line in current_block:
                    if block_line.startswith("TITLE="):
                        m = re.search(r"scan=(\d+)", block_line)
                        if m:
                            scan_number = int(m.group(1))
                        break

                if scan_number in target_scans:
                    fout.writelines(current_block)
                    kept_count += 1

                    if merge_all:
                        with open(merged_output_file, "a", encoding="utf-8") as fall:
                            fall.writelines(current_block)

                current_block = []

            elif inside_block:
                current_block.append(line)

    print(f"Done: kept {kept_count} spectra")
    print(f"Output file: {output_file}")

print("\nall done.")
if merge_all:
    print(f"Combined MGF file: {merged_output_file}")